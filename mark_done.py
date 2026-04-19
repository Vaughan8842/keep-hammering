#!/usr/bin/env python3
"""
Mark tasks as Done (or any status) in Task Tracker.xlsx.

Usage:
  python3 ~/Documents/KeepHammering/mark_done.py "partial task name"
  python3 ~/Documents/KeepHammering/mark_done.py "partial task name" --status "In Progress"
  python3 ~/Documents/KeepHammering/mark_done.py "partial task name" --undo
"""
import sys
import os
import openpyxl

TRACKER = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Work/STM/Task Tracker.xlsx"
)

VALID_STATUSES = ["Not Started", "In Progress", "Done", "Blocked"]


def find_tasks(wb, query):
    results = []
    for sheet_name in ["Work", "Personal"]:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            name = row[0].value
            if name and query.lower() in str(name).lower():
                results.append((sheet_name, row_idx, str(name), row[3].value))
    return results


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    query = sys.argv[1]

    if "--undo" in sys.argv:
        new_status = "Not Started"
    elif "--status" in sys.argv:
        idx = sys.argv.index("--status")
        if idx + 1 >= len(sys.argv):
            print(f"--status requires a value. Valid: {', '.join(VALID_STATUSES)}")
            sys.exit(1)
        new_status = sys.argv[idx + 1]
        if new_status not in VALID_STATUSES:
            print(f"Invalid status '{new_status}'. Valid: {', '.join(VALID_STATUSES)}")
            sys.exit(1)
    else:
        new_status = "Done"

    wb = openpyxl.load_workbook(TRACKER)
    matches = find_tasks(wb, query)

    if not matches:
        print(f"No task found matching: '{query}'")
        sys.exit(1)

    if len(matches) > 1:
        print(f"Multiple tasks match '{query}' — be more specific:")
        for sheet, row_idx, name, status in matches:
            print(f"  [{sheet}] {name!r}  (currently: {status})")
        sys.exit(1)

    sheet, row_idx, name, old_status = matches[0]
    wb[sheet].cell(row=row_idx, column=4).value = new_status
    wb.save(TRACKER)
    print(f"[{sheet}] {name!r}  {old_status} → {new_status}")


if __name__ == "__main__":
    main()
