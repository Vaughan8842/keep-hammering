#!/usr/bin/env python3
"""
Build Task Tracker.xlsx
Creates Dashboard, Work, and Personal sheets with headers, dropdowns,
conditional formatting, and formula-driven Dashboard panels.

Run: python3 ~/Documents/KeepHammering/build_task_tracker.py
Output: ~/Library/CloudStorage/OneDrive-Personal/Work/STM/Task Tracker.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Work/STM/Task Tracker.xlsx"
)

MAX_ROW = 500

C_HEADER_BG = "1A2744"
C_HEADER_FG = "FFFFFF"
C_ALT_ROW   = "F5F7FA"
C_RED       = "FFCCCC"
C_YELLOW    = "FFF9C4"
C_GREEN     = "C8E6C9"
C_BLUE      = "BBDEFB"

TASK_COLS = [
    ("Task Name",          45),
    ("Priority",           10),
    ("Due Date",           12),
    ("Status",             16),
    ("Project / Category", 22),
    ("Notes",              35),
    ("Date Added",         13),
    ("Source",             14),
]

COMPLETED_COLS = [
    ("Task Name",      45),
    ("Date Completed", 14),
    ("Source",         14),
]


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def build_task_sheet(ws):
    ws.row_dimensions[1].height = 22
    for i, (label, width) in enumerate(TASK_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = Font(bold=True, color=C_HEADER_FG, size=11)
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    dv_priority = DataValidation(
        type="list", formula1='"High,Med,Low"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid", error="Choose High, Med, or Low"
    )
    ws.add_data_validation(dv_priority)
    dv_priority.sqref = f"B2:B{MAX_ROW}"

    dv_status = DataValidation(
        type="list", formula1='"Not Started,In Progress,Done,Blocked"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid", error="Choose a valid status"
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = f"D2:D{MAX_ROW}"

    cf = f"A2:H{MAX_ROW}"
    # Done = green (applied first — prevents done tasks showing as overdue)
    ws.conditional_formatting.add(cf, FormulaRule(
        formula=['AND($A2<>"",$D2="Done")'],
        fill=fill(C_GREEN), stopIfTrue=True
    ))
    # Overdue = red
    ws.conditional_formatting.add(cf, FormulaRule(
        formula=['AND($A2<>"",$C2<>"",$C2<TODAY(),$D2<>"Done")'],
        fill=fill(C_RED), stopIfTrue=True
    ))
    # Blocked = blue
    ws.conditional_formatting.add(cf, FormulaRule(
        formula=['AND($A2<>"",$D2="Blocked")'],
        fill=fill(C_BLUE), stopIfTrue=True
    ))
    # Due today = yellow
    ws.conditional_formatting.add(cf, FormulaRule(
        formula=['AND($A2<>"",$C2=TODAY(),$D2<>"Done")'],
        fill=fill(C_YELLOW)
    ))


def build_completed_sheet(ws):
    ws.row_dimensions[1].height = 22
    for i, (label, width) in enumerate(COMPLETED_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = Font(bold=True, color=C_HEADER_FG, size=11)
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def section_hdr(ws, row, text, col_start, col_end, bg):
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = fill(bg)


def col_hdrs(ws, row, pairs, bg="444444"):
    ws.row_dimensions[row].height = 16
    for col, label in pairs:
        cell = ws.cell(row=row, column=col)
        cell.value = label
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def list_row(ws, row, sheet, cond, k):
    base = f"ROW({sheet}!$A$2:$A${MAX_ROW})-ROW({sheet}!$A$2)+1"

    def pull(col):
        idx = f'SMALL(IF({cond},{base}),{k})'
        val = f'INDEX({sheet}!${col}$2:${col}${MAX_ROW},{idx})'
        return f'=IFERROR(IF({val}=0,"",{val}),"")'

    ws.cell(row=row, column=2).value = ArrayFormula(f"B{row}", pull("A"))
    due_cell = ws.cell(row=row, column=3)
    due_cell.value = ArrayFormula(f"C{row}", pull("C"))
    due_cell.number_format = "MM/DD"
    ws.cell(row=row, column=4).value = ArrayFormula(f"D{row}", pull("B"))
    ws.cell(row=row, column=5).value = f'=IF(B{row}<>"","{sheet}","")'

    shade = C_ALT_ROW if row % 2 == 0 else "FFFFFF"
    for c in range(2, 6):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(shade)
        cell.font = Font(size=10)
        cell.alignment = Alignment(vertical="center")


def cond_overdue(s):
    return (f'({s}!$A$2:$A${MAX_ROW}<>"")*({s}!$C$2:$C${MAX_ROW}<>"")'
            f'*({s}!$C$2:$C${MAX_ROW}<TODAY())*({s}!$D$2:$D${MAX_ROW}<>"Done")')


def cond_today(s):
    return (f'({s}!$A$2:$A${MAX_ROW}<>"")'
            f'*({s}!$C$2:$C${MAX_ROW}=TODAY())*({s}!$D$2:$D${MAX_ROW}<>"Done")')


def cond_hipri(s):
    return (f'({s}!$A$2:$A${MAX_ROW}<>"")*({s}!$B$2:$B${MAX_ROW}="High")'
            f'*({s}!$D$2:$D${MAX_ROW}<>"Done")')


def cond_open(s):
    return (f'({s}!$A$2:$A${MAX_ROW}<>"")'
            f'*({s}!$D$2:$D${MAX_ROW}<>"Done")')


def build_dashboard(ws):
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 3
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 8

    # Row 1: title
    ws.row_dimensions[1].height = 32
    ws.merge_cells('B1:J1')
    t = ws['B1']
    t.value = "TASK TRACKER — DASHBOARD"
    t.font = Font(bold=True, size=16, color=C_HEADER_FG)
    t.fill = fill(C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: date
    ws.merge_cells('B2:J2')
    d = ws['B2']
    d.value = '=TEXT(TODAY(),"MMMM D, YYYY")'
    d.font = Font(italic=True, color="888888", size=10)
    d.alignment = Alignment(horizontal="center")

    ws.row_dimensions[3].height = 6

    # Rows 4-5: summary cards
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28

    cards = [
        ("TOTAL OPEN",
         f'=COUNTIFS(Work!D$2:D${MAX_ROW},"<>Done",Work!A$2:A${MAX_ROW},"<>")'
         f'+COUNTIFS(Personal!D$2:D${MAX_ROW},"<>Done",Personal!A$2:A${MAX_ROW},"<>")',
         "3A5A8C"),
        ("OVERDUE",
         f'=COUNTIFS(Work!C$2:C${MAX_ROW},"<"&TODAY(),Work!D$2:D${MAX_ROW},"<>Done",'
         f'Work!A$2:A${MAX_ROW},"<>",Work!C$2:C${MAX_ROW},"<>")'
         f'+COUNTIFS(Personal!C$2:C${MAX_ROW},"<"&TODAY(),Personal!D$2:D${MAX_ROW},"<>Done",'
         f'Personal!A$2:A${MAX_ROW},"<>",Personal!C$2:C${MAX_ROW},"<>")',
         "C62828"),
        ("DUE TODAY",
         f'=COUNTIFS(Work!C$2:C${MAX_ROW},TODAY(),Work!D$2:D${MAX_ROW},"<>Done",'
         f'Work!A$2:A${MAX_ROW},"<>")'
         f'+COUNTIFS(Personal!C$2:C${MAX_ROW},TODAY(),Personal!D$2:D${MAX_ROW},"<>Done",'
         f'Personal!A$2:A${MAX_ROW},"<>")',
         "E65100"),
        ("BLOCKED",
         f'=COUNTIFS(Work!D$2:D${MAX_ROW},"Blocked",Work!A$2:A${MAX_ROW},"<>")'
         f'+COUNTIFS(Personal!D$2:D${MAX_ROW},"Blocked",Personal!A$2:A${MAX_ROW},"<>")',
         "1565C0"),
        ("TOTAL DONE",
         f'=COUNTIF(Completed!B$2:B${MAX_ROW},TODAY())',
         "2E7D32"),
    ]

    for i, (label, formula, color) in enumerate(cards):
        col = 2 + i * 2
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        lbl = ws.cell(row=4, column=col)
        lbl.value = label
        lbl.font = Font(bold=True, color="FFFFFF", size=8)
        lbl.fill = fill(color)
        lbl.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        val = ws.cell(row=5, column=col)
        val.value = formula
        val.font = Font(bold=True, color="FFFFFF", size=18)
        val.fill = fill(color)
        val.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[6].height = 8

    # ── ALL OPEN TASKS (rows 7–31) ───────────────────────────────────────────────
    section_hdr(ws, 7, "📋  ALL OPEN TASKS", 2, 5, "1A2744")

    section_hdr(ws, 8, "  WORK", 2, 5, "3A5A8C")
    col_hdrs(ws, 9, [(2, "Task Name"), (3, "Due"), (4, "Priority"), (5, "Sheet")])
    r = 10
    for k in range(1, 11):
        list_row(ws, r, "Work", cond_open("Work"), k)
        r += 1

    section_hdr(ws, 20, "  PERSONAL", 2, 5, "2E7D32")
    col_hdrs(ws, 21, [(2, "Task Name"), (3, "Due"), (4, "Priority"), (5, "Sheet")])
    r = 22
    for k in range(1, 11):
        list_row(ws, r, "Personal", cond_open("Personal"), k)
        r += 1

    # ── OVERDUE (rows 32–42, left) + STATUS BREAKDOWN (rows 32–38, right) ────────
    section_hdr(ws, 32, "🔴  OVERDUE", 2, 5, "C62828")
    col_hdrs(ws, 33, [(2, "Task Name"), (3, "Due"), (4, "Priority"), (5, "Sheet")])

    r = 34
    for sheet in ["Work", "Personal"]:
        for k in range(1, 5):
            list_row(ws, r, sheet, cond_overdue(sheet), k)
            r += 1

    section_hdr(ws, 32, "📊  STATUS BREAKDOWN", 7, 10, "2D4A6B")
    col_hdrs(ws, 33, [(7, "Status"), (8, "Work"), (9, "Personal"), (10, "Total")])

    for i, status in enumerate(["Not Started", "In Progress", "Done", "Blocked"]):
        row = 34 + i
        ws.row_dimensions[row].height = 16
        shade = C_ALT_ROW if row % 2 == 0 else "FFFFFF"
        ws.cell(row=row, column=7).value = status
        ws.cell(row=row, column=8).value = (
            f'=COUNTIFS(Work!D$2:D${MAX_ROW},"{status}",Work!A$2:A${MAX_ROW},"<>")'
        )
        ws.cell(row=row, column=9).value = (
            f'=COUNTIFS(Personal!D$2:D${MAX_ROW},"{status}",Personal!A$2:A${MAX_ROW},"<>")'
        )
        ws.cell(row=row, column=10).value = f"=H{row}+I{row}"
        for c in range(7, 11):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill(shade)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=7).alignment = Alignment(
            horizontal="left", vertical="center", indent=1
        )

    # ── DUE TODAY (rows 44–52) ───────────────────────────────────────────────────
    ws.row_dimensions[43].height = 6
    section_hdr(ws, 44, "🟡  DUE TODAY", 2, 5, "E65100")
    col_hdrs(ws, 45, [(2, "Task Name"), (3, "Due"), (4, "Priority"), (5, "Sheet")])

    r = 46
    for sheet in ["Work", "Personal"]:
        for k in range(1, 4):
            list_row(ws, r, sheet, cond_today(sheet), k)
            r += 1

    # ── HIGH PRIORITY (rows 54–64) ───────────────────────────────────────────────
    ws.row_dimensions[53].height = 6
    section_hdr(ws, 54, "🔥  HIGH PRIORITY (OPEN)", 2, 5, "4A148C")
    col_hdrs(ws, 55, [(2, "Task Name"), (3, "Due"), (4, "Priority"), (5, "Sheet")])

    r = 56
    for sheet in ["Work", "Personal"]:
        for k in range(1, 5):
            list_row(ws, r, sheet, cond_hipri(sheet), k)
            r += 1


def main():
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_work = wb.create_sheet("Work")
    ws_personal = wb.create_sheet("Personal")
    ws_completed = wb.create_sheet("Completed")

    build_task_sheet(ws_work)
    build_task_sheet(ws_personal)
    build_completed_sheet(ws_completed)
    build_dashboard(ws_dash)

    output = os.path.expanduser(OUTPUT_PATH)
    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    print(f"Saved: {output}")


if __name__ == "__main__":
    main()
