#!/usr/bin/env python3
"""
Keep Hammering — Task Tracker Sync
Reads today's Apple Note, extracts TODO items from 'Today's TODO' (work)
and 'Personal TODO' (personal) checklist sections, and appends new tasks
to Task Tracker.xlsx. Safe to run multiple times — deduplicates by
task name + date added.

Run:      python3 ~/Documents/KeepHammering/sync_tasks.py
Auto-run: Mac Shortcuts, daily at 6 AM
"""

import subprocess
import re
import os
from datetime import datetime, timedelta
import openpyxl

TRACKER_PATH = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Work/STM/Task Tracker.xlsx"
)

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
DAY_NAMES = [
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
]


def yesterday_note_title():
    yesterday = datetime.today() - timedelta(days=1)
    return f"{DAY_NAMES[yesterday.weekday()]}, {yesterday.day} {MONTH_NAMES[yesterday.month - 1]} {yesterday.year}"


def get_note_html(note_title):
    safe = note_title.replace('"', '\\"')
    script = f'''
tell application "Notes"
    tell default account
        set stmFolder to missing value
        repeat with f in folders
            if name of f is "STM" then
                set stmFolder to f
                exit repeat
            end if
        end repeat
        if stmFolder is missing value then return ""
        set targetFolder to missing value
        repeat with sf in folders of stmFolder
            if name of sf is "Daily Notes" then
                set targetFolder to sf
                exit repeat
            end if
        end repeat
        if targetFolder is missing value then return ""
        repeat with n in notes of targetFolder
            if name of n is "{safe}" then
                return body of n
            end if
        end repeat
    end tell
end tell
return ""
'''
    result = subprocess.run(
        ["osascript", "-e", script], capture_output=True, text=True, timeout=30
    )
    return result.stdout.strip()


def parse_todos(html_body):
    """
    Returns (work_tasks, personal_tasks).
    Each task is (task_name: str, due_date: date | None).
    Handles both <ul class="checklist"> (new notes) and plain <ul> (older notes).
    For checklists, checked items are skipped. For plain lists, all non-empty items
    are included.
    """
    # Strip embedded image data so it doesn't interfere with parsing
    html = re.sub(r'src="data:[^"]*"', 'src=""', html_body)

    def items_after_header(header_pattern):
        m = re.search(header_pattern, html, re.IGNORECASE)
        if not m:
            return []
        after = html[m.end():]
        ul = re.search(r'<ul([^>]*)>(.*?)</ul>', after, re.DOTALL)
        if not ul:
            return []
        is_checklist = 'checklist' in ul.group(1).lower()
        found = []
        for li in re.finditer(r'<li([^>]*)>(.*?)</li>', ul.group(2), re.DOTALL):
            if is_checklist and 'checked' in li.group(1).lower():
                continue
            text = re.sub(r'<[^>]+>', '', li.group(2)).strip()
            if text:
                found.append(text)
        return found

    # Match both straight apostrophe (') and smart apostrophe (\u2019)
    work_items = items_after_header(r"Today['\u2019]s\s+TODO")
    personal_items = items_after_header(r"Personal\s+TODO")

    return (
        [_parse_due(i) for i in work_items],
        [_parse_due(i) for i in personal_items],
    )


def _parse_due(text):
    """Strip mm/dd tag from task text, return (clean_name, date_or_None)."""
    match = re.search(r'\b(\d{1,2}/\d{1,2})\b', text)
    if not match:
        return text.strip(), None

    raw = match.group(1)
    clean = (text[: match.start()] + text[match.end() :]).strip()

    month, day = int(raw.split('/')[0]), int(raw.split('/')[1])
    today = datetime.today()
    try:
        candidate = datetime(today.year, month, day).date()
        # If more than 60 days in the past, assume next year
        if (today.date() - candidate).days > 60:
            candidate = datetime(today.year + 1, month, day).date()
        return clean, candidate
    except ValueError:
        return text.strip(), None


def existing_keys(ws):
    """Return set of (task_name_lower, date_str) already in sheet."""
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        date_added = row[6]
        if name:
            if hasattr(date_added, 'date'):
                ds = str(date_added.date())
            else:
                ds = str(date_added or '')
            keys.add((str(name).strip().lower(), ds))
    return keys


def append_tasks(ws, tasks, today_dt):
    today_str = str(today_dt.date())
    keys = existing_keys(ws)
    added = 0
    for name, due_date in tasks:
        if not name:
            continue
        if (name.strip().lower(), today_str) in keys:
            continue
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=2).value = ""              # Priority — user fills in
        if due_date:
            ws.cell(row=r, column=3).value = datetime(
                due_date.year, due_date.month, due_date.day
            )
            ws.cell(row=r, column=3).number_format = "MM/DD/YY"
        ws.cell(row=r, column=4).value = "Not Started"
        ws.cell(row=r, column=5).value = ""              # Category — user fills in
        ws.cell(row=r, column=6).value = ""              # Notes — user fills in
        ws.cell(row=r, column=7).value = today_dt
        ws.cell(row=r, column=7).number_format = "MM/DD/YY"
        ws.cell(row=r, column=8).value = "Apple Notes"
        added += 1
    return added


def load_daily_tasks(wb):
    """
    Read the Daily Tasks sheet and return (work_tasks, personal_tasks).
    Each entry is (task_name, priority_or_None) — no due date for daily tasks.
    """
    if "Daily Tasks" not in wb.sheetnames:
        return [], []
    ws = wb["Daily Tasks"]
    work, personal = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, sheet, priority = (row[0], row[1], row[2]) if len(row) >= 3 else (row[0], row[1] if len(row) > 1 else None, None)
        if not name:
            continue
        entry = (str(name).strip(), priority)
        if str(sheet or "").strip().lower() == "personal":
            personal.append(entry)
        else:
            work.append(entry)
    return work, personal


def append_daily_tasks(ws, daily_tasks, run_date):
    """Append daily tasks that aren't already present for run_date."""
    date_str = str(run_date.date() if hasattr(run_date, 'date') else run_date)
    keys = existing_keys(ws)
    added = 0
    for name, priority in daily_tasks:
        if not name:
            continue
        if (name.strip().lower(), date_str) in keys:
            continue
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=2).value = priority or ""
        ws.cell(row=r, column=4).value = "Not Started"
        ws.cell(row=r, column=7).value = run_date
        ws.cell(row=r, column=7).number_format = "MM/DD/YY"
        ws.cell(row=r, column=8).value = "Daily"
        added += 1
    return added


def archive_done_tasks(wb, today):
    """Move Done rows from Work/Personal to Completed sheet. Returns count archived."""
    if "Completed" not in wb.sheetnames:
        return 0
    ws_completed = wb["Completed"]
    archived = 0
    for sheet_name in ["Work", "Personal"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_to_delete = []
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            status = ws.cell(row=row_idx, column=4).value
            if name and str(status or "").strip().lower() == "done":
                source = ws.cell(row=row_idx, column=8).value or "Manual"
                r = ws_completed.max_row + 1
                ws_completed.cell(row=r, column=1).value = name
                ws_completed.cell(row=r, column=2).value = today
                ws_completed.cell(row=r, column=2).number_format = "MM/DD/YY"
                ws_completed.cell(row=r, column=3).value = source
                rows_to_delete.append(row_idx)
                archived += 1
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
    return archived


def main():
    title = yesterday_note_title()
    print(f"Reading: {title}")

    tracker = os.path.expanduser(TRACKER_PATH)
    if not os.path.exists(tracker):
        print(f"Task Tracker not found at {TRACKER_PATH}")
        print("Run build_task_tracker.py first.")
        return

    wb = openpyxl.load_workbook(tracker)
    today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)

    archived = archive_done_tasks(wb, today)
    if archived:
        print(f"Archived {archived} completed task(s).")

    # Daily recurring tasks (always added, regardless of Apple Notes)
    daily_work, daily_personal = load_daily_tasks(wb)
    dw = append_daily_tasks(wb["Work"], daily_work, yesterday)
    dp = append_daily_tasks(wb["Personal"], daily_personal, yesterday)
    print(f"Daily tasks: added {dw} work, {dp} personal")

    # Tasks from Apple Notes
    html = get_note_html(title)
    if not html:
        print("Note not found — skipping Apple Notes sync.")
    else:
        work_tasks, personal_tasks = parse_todos(html)
        print(f"Notes parsed: {len(work_tasks)} work, {len(personal_tasks)} personal")
        w = append_tasks(wb["Work"], work_tasks, yesterday)
        p = append_tasks(wb["Personal"], personal_tasks, yesterday)
        print(f"Notes added: {w} work, {p} personal")

    wb.save(tracker)
    print("Saved.")


if __name__ == "__main__":
    main()
